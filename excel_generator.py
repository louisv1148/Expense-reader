import pandas as pd
from database import ExpenseDatabase
from datetime import datetime
import os

class ExcelExpenseGenerator:
    def __init__(self):
        self.db = ExpenseDatabase()
    
    def generate_monthly_report(self, output_filename=None):
        """Generate Excel report matching your company format"""
        
        # Get all reviewed receipts
        receipts = [r for r in self.db.get_all_receipts() if r['reviewed']]
        
        if not receipts:
            raise ValueError("No reviewed receipts found. Please review receipts in the web interface first.")
        
        # Prepare data in the required format
        data = []
        for receipt in receipts:
            # Calculate USD amount (receipts are in MXN, convert to USD)
            amount_usd = receipt.get('amount_mxn')  # This field actually stores USD
            if not amount_usd and receipt.get('total_amount') and receipt.get('fx_rate'):
                markup = receipt.get('markup_percent', 2.5)
                # total_amount is in MXN, convert to USD with markup
                usd_base = receipt['total_amount'] / receipt['fx_rate']
                amount_usd = usd_base * (1 + markup / 100)
            
            # Format date for filename if available
            formatted_filename = receipt['filename']
            if receipt.get('date') and receipt.get('restaurant_name'):
                from filename_utils import format_receipt_filename
                new_filename = format_receipt_filename(receipt['date'], receipt['restaurant_name'])
                if new_filename:
                    formatted_filename = new_filename
            
            # Format date as DD/MM/YYYY
            formatted_date = ''
            if receipt.get('date'):
                try:
                    date_obj = datetime.strptime(receipt['date'], '%Y-%m-%d')
                    formatted_date = date_obj.strftime('%d/%m/%Y')
                except:
                    formatted_date = receipt['date']  # Fallback to original format
            
            row = {
                'Fecha': formatted_date,
                'Proveedor': receipt.get('restaurant_name', ''),
                'Detalle': receipt.get('detalle', ''),
                'Monto': round(amount_usd, 2) if amount_usd else 0,  # USD amount
                'Reembolso': receipt.get('reembolso', ''),
                'Cuenta contable': receipt.get('cuenta_contable', 'Comidas con Clientes'),
                'Pais': receipt.get('pais', 'MX'),
                'CC': receipt.get('cc', 'Alternativos')
            }
            data.append(row)
        
        # Create DataFrame with specific column order
        column_order = ['Fecha', 'Proveedor', 'Detalle', 'Monto', 'Reembolso', 'Cuenta contable', 'Pais', 'CC']
        df = pd.DataFrame(data, columns=column_order)
        
        # Generate filename if not provided
        if not output_filename:
            current_date = datetime.now()
            output_filename = f"{current_date.year}_{current_date.month:02d}_Gastos_MX.xlsx"
        
        # Write to Excel with formatting
        with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Gastos', index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Gastos']
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Add some basic formatting
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Header formatting
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            # Number formatting for amounts
            for row in worksheet.iter_rows(min_row=2):
                # Monto column (column F)
                if row[5].value and isinstance(row[5].value, (int, float)):
                    row[5].number_format = '#,##0.00'
                # Monto USD column (column G)
                if row[6].value and isinstance(row[6].value, (int, float)):
                    row[6].number_format = '#,##0.00'
                # FX Rate column (column H)
                if row[7].value and isinstance(row[7].value, (int, float)):
                    row[7].number_format = '0.00'
        
        return output_filename
    
    def generate_summary_stats(self):
        """Generate summary statistics for the report"""
        receipts = [r for r in self.db.get_all_receipts() if r['reviewed']]
        
        if not receipts:
            return None
        
        # Calculate totals by category
        category_totals = {}
        cc_totals = {}
        total_mxn = 0
        total_usd = 0
        
        for receipt in receipts:
            # Calculate MXN amount
            amount_mxn = receipt.get('amount_mxn', 0)
            if not amount_mxn and receipt.get('total_amount') and receipt.get('fx_rate'):
                markup = receipt.get('markup_percent', 2.5)
                amount_with_markup = receipt['total_amount'] * (1 + markup / 100)
                amount_mxn = amount_with_markup * receipt['fx_rate']
            
            # Totals by category
            category = receipt.get('cuenta_contable', 'Unknown')
            category_totals[category] = category_totals.get(category, 0) + amount_mxn
            
            # Totals by cost center
            cc = receipt.get('cc', 'Unknown')
            cc_totals[cc] = cc_totals.get(cc, 0) + amount_mxn
            
            total_mxn += amount_mxn
            total_usd += receipt.get('total_amount', 0)
        
        return {
            'total_receipts': len(receipts),
            'total_mxn': total_mxn,
            'total_usd': total_usd,
            'category_totals': category_totals,
            'cc_totals': cc_totals,
            'avg_fx_rate': sum(r.get('fx_rate', 0) for r in receipts) / len(receipts) if receipts else 0
        }

def main():
    """Test Excel generation"""
    generator = ExcelExpenseGenerator()
    try:
        filename = generator.generate_monthly_report()
        print(f"Excel report generated successfully: {filename}")
        
        stats = generator.generate_summary_stats()
        if stats:
            print(f"\nSummary:")
            print(f"Total receipts: {stats['total_receipts']}")
            print(f"Total MXN: ${stats['total_mxn']:,.2f}")
            print(f"Total USD: ${stats['total_usd']:,.2f}")
            print(f"Average FX Rate: {stats['avg_fx_rate']:.2f}")
            
    except Exception as e:
        print(f"Error generating Excel: {e}")

if __name__ == "__main__":
    main()